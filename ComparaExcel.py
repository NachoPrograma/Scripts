import glob
from openpyxl import Workbook
import os

# Crear un nuevo archivo de Excel
libro = Workbook()
hoja = libro.active
fila = 0

# Especificar los nombres de columna personalizados

hoja.cell(row=1, column=1, value="Hospital")
hoja.cell(row=1, column=2, value="Fecha")
hoja.cell(row=1, column=3, value="Almacen")
hoja.cell(row=1, column=4, value="Codigo")
hoja.cell(row=1, column=5, value="Denominacion")
hoja.cell(row=1, column=6, value="Cantidad FT")
hoja.cell(row=1, column=7, value="Cantidad APD")
hoja.cell(row=1, column=8, value="Diferencia")

#ruta_principal = input("Por favor, introduce la ruta principal: ")
ruta_principal = r"C:\Users\Admin\Desktop\Hospitales_Compara"

# Obtener la lista de carpetas en la ruta principal
carpetas = glob.glob(os.path.join(ruta_principal, "*"))

columna_inicial = 2  # Columna inicial para el primer archivo
row = 2
for carpeta in carpetas:
    if not os.path.isdir(carpeta):
        continue  # Saltar si no es una carpeta válida

    # Obtener el nombre de la carpeta
    nombre_carpeta = os.path.basename(carpeta)
    # Obtener la lista de archivos TXT en la carpeta
    archivos_txt = glob.glob(os.path.join(carpeta, "*.txt"))
    # Ordenar los archivos por fecha de modificación (de más reciente a más antiguo)
    archivos_recientes = sorted(archivos_txt, key=os.path.getmtime, reverse=True)
    # Tomar los 7 archivos más recientes
    archivos_seleccionados = archivos_recientes[:10]   

    for i, archivo_txt in enumerate(archivos_seleccionados):
        nombre_archivo = os.path.basename(archivo_txt)  # Obtener el nombre del archivo
        fecha_archivo = nombre_archivo.split('_')[-1].split('.')[0]  # Extraer la fecha del nombre del archivo
        anio = fecha_archivo[:4]
        mes = fecha_archivo[4:6]
        dia = fecha_archivo[6:8]
        fecha_formateada = f"{mes}-{dia}"
        fecha_inicio = f"{anio}/{mes}/{dia}"     

        with open(archivo_txt, 'r') as archivo:
            lineas = archivo.readlines()           
            # Reiniciar la variable 'row' para cada archivo
            for linea in lineas:
                # Utilizar expresiones regulares para extraer los datos
                datos = linea.split('\t')

                if len(datos):                  
                    almacen = datos[1]
                    if almacen.strip().lower() != "no hay mas descuadres.":
                        codigo = datos[2]
                        denominacion = datos[3]
                        cantidadFTespacios = datos[5]
                        cantidadFT = "".join(cantidadFTespacios.split())
                        cantidadAPDespacios = datos[7]
                        cantidadAPD = "".join(cantidadAPDespacios.split())
                        if cantidadFT == 'Noexiste' or cantidadAPD == 'Noexiste':
                            diferencia = 'Null'
                        else:
                            diferencia = int(cantidadFT) - int(cantidadAPD)

                        # Agregar los datos a las columnas correspondientes
                        hoja.cell(row, column=1, value=nombre_carpeta)
                        hoja.cell(row, column=2, value=fecha_inicio)
                        hoja.cell(row, column=columna_inicial + 1, value=almacen)
                        hoja.cell(row, column=columna_inicial + 2, value=codigo)
                        hoja.cell(row, column=columna_inicial + 3, value=denominacion)
                        hoja.cell(row, column=columna_inicial + 4, value=cantidadFT)
                        hoja.cell(row, column=columna_inicial + 5, value=cantidadAPD)
                        hoja.cell(row, column=columna_inicial + 6, value=diferencia)

                        # Incrementar la variable 'row' para la siguiente fila
                        row += 1
                    
# Guardar el archivo de Excel
libro.save('DatosCompar.xlsx')
input("Finalizado. Pulsa cualquier tecla para salir de la terminal.")